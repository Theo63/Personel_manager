import datetime
import calendar
from itertools import chain
import collections
import xlrd
from openpyxl import Workbook
import string

year=int(input('Έτος: '))
month=int(input('Μήνας: '))
x=None

catt={}
weeks=calendar.monthcalendar(year,month)

categ4=input('Αργίες για τον μήνα: ')
categ4=[int(x) for x in categ4.split(',') if x.strip().isdigit()]

for week in weeks:
    i=1
    for day in week:
        if day is 0:
            i=i+1
            pass
        elif day in categ4:
            catt[day]=[]
            catt[day].append('cat4')
            i=i+1
        else:
            if i<5:
                catt[day]=[]
                catt[day].append('cat1')
                i=i+1
            elif i==5:
                catt[day]=[]
                catt[day].append('cat2')
                i=i+1
            elif i==6:
                catt[day]=[]
                catt[day].append('cat3')
                i=i+1
            else:
                catt[day]=[]
                catt[day].append('cat4')
                i=i+1

# Διάβασμα ονομάτων.
loc=('Duty_sheet '+str(month-1)+'.xlsx')

# To open Workbook
wb = xlrd.open_workbook(loc)
sheet = wb.sheet_by_index(1)
num_rows=sheet.nrows-1
num_cols=sheet.ncols-1

book = Workbook()
sheet1 = book.active        #opens sheet to write duty dates
sheet2 = book.create_sheet()        #opens sheet to write duty control numbers

audm_list=[]
for k in range(sheet.nrows):        #διαβαζει απο το εξελ την καθε γραμμη και τα βαζει σε μια λιστα σαν λιστες
        audm_list.append(sheet.row_values(k))

audm_dic={}
value2=[]
for inner_l in audm_list:
    key=inner_l[0:1:] #παιρνει το ονομα του καθε ΑΥΔΜ απο την λιστα σαν μια λιστα με ενα στοιχειο
    key1=key[0]    #παιρνει απο την λιστα το στοιχειο για να γινει key (σαν λιστα δεν γινεται )
    value=inner_l[1:5:] #παιρνει απο την λιστα τα υπολοιπα στοιχεια σαν λιστα
    value = [ int(x) for x in value ] #κανει ιντ τα φλοατσ
    value2=inner_l[5]
    value2=[int(x) for x in value2.split(',') if x.strip().isdigit()]
    audm_dic[key1]=value #τα βαζει στο λεξικο audm_dic
    audm_dic[key1].append(value2)

for name in audm_dic.items():
    audm_dic[name[0]]={}
    audm_dic[name[0]]['cat1']=name[1][0]
    audm_dic[name[0]]['cat2']=name[1][1]
    audm_dic[name[0]]['cat3']=name[1][2]
    audm_dic[name[0]]['cat4']=name[1][3]
    audm_dic[name[0]]['except_dates']=name[1][4]
    audm_dic[name[0]]['except1']=False
    audm_dic[name[0]]['total']=0


tdays_bfr=None
day_bfr=None
for day in catt.items():
    # print(tdays_bfr,day_bfr)

##CAT 1
    minimum1=audm_dic['Papadimitriou']['cat1']
    minimum2=audm_dic['Papadimitriou']['cat2']
    minimum3=audm_dic['Papadimitriou']['cat3']
    minimum4=audm_dic['Papadimitriou']['cat4']

    if day[1][0]=="cat1":
        m=[]
        for audm in audm_dic.items():
            m.append(audm[1]['cat1'])
            minimum1=min(m)

        for audm in audm_dic.items():                                   ##check το cat1 του κάθε αύδμ αν είναι μικρότερο από το minimum
            if audm[0] is not day_bfr and audm[0] is not tdays_bfr:     ##check τις προηγούμενες μέρες αν ειναι ο αυδμ υπηρεσία
                if audm_dic[audm[0]]['total']<3 and day[0] not in audm_dic[audm[0]]['except_dates']:                        ##check αν εχει μέχρι τρείς υπηρεσίες
                    if audm[1]['cat1']<=minimum1:                            ##check αν είναι στο μίνιμουμ υπηρεσιων στην κατήγορία καθημερινές
                        catt[day[0]].append(audm[0])                    ##προσθεση του ονονόματός στο catt
                        audm_dic[audm[0]]['cat1']=audm_dic[audm[0]]['cat1']+1   ##και +1 στο cat1 του
                        audm_dic[audm[0]]['total']=audm_dic[audm[0]]['total']+1
                        break
                    else:
                        pass

##CAT 2

    elif day[1][0]=="cat2":
        for audm in audm_dic.items():       ##minimum() audm[0]=name audm[1]=dictionary ['key']
            m=[]
            for audm in audm_dic.items():
                m.append(audm[1]['cat2'])
                minimum2=min(m)


        for audm in audm_dic.items():       ##check το cat1 του κάθε αύδμ αν είναι μικρότερο από το minimum
            if audm[0] is not day_bfr and audm[0] is not tdays_bfr:
                if audm_dic[audm[0]]['total']<3 and audm_dic[audm[0]]['except1']==False and day[0] not in audm_dic[audm[0]]['except_dates']:
                    if audm[1]['cat2']<=minimum2:
                        catt[day[0]].append(audm[0])    ##προσθεση του ονονόματός στο catt
                        audm_dic[audm[0]]['cat2']=audm_dic[audm[0]]['cat2']+1   ##και +1 στο cat1 του
                        audm_dic[audm[0]]['total']=audm_dic[audm[0]]['total']+1
                        audm_dic[audm[0]]['except1']=True
                        break
                    else:
                        pass

##CAT 3

    elif day[1][0]=="cat3":
        for audm in audm_dic.items():       ##minimum() audm[0]=name audm[1]=dictionary ['key']
            m=[]
            for audm in audm_dic.items():
                m.append(audm[1]['cat3'])
                minimum3=min(m)


        for audm in audm_dic.items():       ##check το cat1 του κάθε αύδμ αν είναι μικρότερο από το minimum
            if audm[0] is not day_bfr and audm[0] is not tdays_bfr:
                if audm_dic[audm[0]]['total']<3 and audm_dic[audm[0]]['except1']==False and day[0] not in audm_dic[audm[0]]['except_dates']:
                    if audm[1]['cat3']<=minimum3:
                        catt[day[0]].append(audm[0])    ##προσθεση του ονονόματός στο catt
                        audm_dic[audm[0]]['cat3']=audm_dic[audm[0]]['cat3']+1   ##και +1 στο cat1 του
                        audm_dic[audm[0]]['total']=audm_dic[audm[0]]['total']+1
                        audm_dic[audm[0]]['except1']=True
                        break
                    else:
                        pass
                else:
                    pass

##CAT 4

    elif day[1][0]=="cat4":
        for audm in audm_dic.items():       ##minimum() audm[0]=name audm[1]=dictionary ['key']
            m=[]
            for audm in audm_dic.items():
                m.append(audm[1]['cat4'])
                minimum4=min(m)


        for audm in audm_dic.items():       ##check το cat1 του κάθε αύδμ αν είναι μικρότερο από το minimum
            if audm[0] is not day_bfr and audm[0] is not tdays_bfr:
                if audm_dic[audm[0]]['total']<3 and audm_dic[audm[0]]['except1']==False and day[0] not in audm_dic[audm[0]]['except_dates']:
                    if audm[1]['cat4']<=minimum4:
                        catt[day[0]].append(audm[0])    ##προσθεση του ονονόματός στο catt
                        audm_dic[audm[0]]['cat4']=audm_dic[audm[0]]['cat4']+1   ##και +1 στο cat1 του
                        audm_dic[audm[0]]['total']=audm_dic[audm[0]]['total']+1
                        audm_dic[audm[0]]['except1']=True
                        break
                    else:
                        pass

    tdays_bfr=day_bfr
    day_bfr=audm[0]

r=1
for row in catt.items(): ##se kathe grammh r sthn stili 1 grafei to key kai sthn stili2 grafei to value
    try:
        sheet1.cell(row=r, column=1).value = r
        sheet1.cell(row=r, column=2, value = row[1][1])
        r=r+1
    except:
        'Some error in name assignment'
# for row in catt.items():
#     print(row[1])

r=1
for row in audm_dic:
    sheet2.cell(row=r, column=1).value = row
    sheet2.cell(row=r, column=2, value = str(int(audm_dic[row]['cat1'])))
    sheet2.cell(row=r, column=3, value = str(int(audm_dic[row]['cat2'])))
    sheet2.cell(row=r, column=4, value = str(int(audm_dic[row]['cat3'])))
    sheet2.cell(row=r, column=5, value = str(int(audm_dic[row]['cat4'])))
    sheet2.cell(row=r, column=6, value = '0,0,0'.translate({ord(i):None for i in "'"}))
    r=r+1

book.save('Duty_sheet '+str(month)+'.xlsx')

## this a new feature of branch 
